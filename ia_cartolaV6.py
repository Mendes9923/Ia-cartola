import requests
import pandas as pd
import random
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

# URL da API do Cartola FC
url = "https://api.cartolafc.globo.com/atletas/mercado"
response = requests.get(url)
dados_mercado = response.json()

# Processamento dos dados dos jogadores
jogadores = dados_mercado['atletas']
clubes = dados_mercado['clubes']
posicoes = dados_mercado['posicoes']

dados = []
for jogador in jogadores:
    clube_id = jogador.get('clube_id')
    clube_nome = clubes.get(str(clube_id), {}).get('nome', 'Desconhecido')
    posicao_id = jogador.get('posicao_id')
    posicao_nome = posicoes.get(str(posicao_id), {}).get('nome', 'Desconhecido')

    scout = jogador.get('scout', {})
    desarmes = scout.get('DS', 0)
    finalizacoes_defendidas = scout.get('FD', 0)
    finalizacoes_fora = scout.get('FF', 0)
    faltas_cometidas = scout.get('FC', 0)

    dados.append({
        "Apelido": jogador.get('apelido', 'Desconhecido'),
        "Clube": clube_nome,
        "Posição": posicao_nome,
        "Preço (C$)": jogador.get('preco_num', 0),
        "Média Pontuação": jogador.get('media_num', 0),
        "Última Pontuação": jogador.get('pontos_num', 0),
        "Mínimo para Valorizar": jogador.get('minimo_para_valorizar', 0),
        "Variação de Preço": jogador.get('variacao_num', 0),
        "Desarmes (DS)": desarmes,
        "Finalizações Defendidas (FD)": finalizacoes_defendidas,
        "Finalizações Fora (FF)": finalizacoes_fora,
        "Faltas Cometidas (FC)": faltas_cometidas
    })

df = pd.DataFrame(dados)

# Criar uma pontuação composta (score personalizado)
df["Score"] = (
    df["Média Pontuação"] * 0.5 +
    df["Desarmes (DS)"] * 0.2 +
    df["Finalizações Defendidas (FD)"] * 0.15 +
    df["Finalizações Fora (FF)"] * 0.1 +
    df["Variação de Preço"] * 0.05
)

# Ordenar os jogadores pelo novo score
df = df.sort_values(by="Score", ascending=False)

# Filtrar jogadores sem faltas cometidas
df = df[df["Faltas Cometidas (FC)"] == 0]

# Entrada do usuário para orçamento
saldo = float(input("Digite o número de cartoletas disponíveis: "))

def criar_time(df, saldo):
    while True:
        time = {
            "Goleiro": random.choice(df[df["Posição"] == "Goleiro"].head(5).to_dict('records')),
            "Zagueiros": random.sample(df[df["Posição"] == "Zagueiro"].head(10).to_dict('records'), 2),
            "Laterais": random.sample(df[df["Posição"] == "Lateral"].head(10).to_dict('records'), 2),
            "Meias": random.sample(df[df["Posição"] == "Meia"].head(15).to_dict('records'), 3),
            "Atacantes": random.sample(df[df["Posição"] == "Atacante"].head(10).to_dict('records'), 3),
            "Técnico": random.choice(df[df["Posição"] == "Técnico"].head(5).to_dict('records'))
        }

        custo_total = (time["Goleiro"]["Preço (C$)"] + sum(j["Preço (C$)"] for j in time["Zagueiros"]) +
                       sum(j["Preço (C$)"] for j in time["Laterais"]) + sum(j["Preço (C$)"] for j in time["Meias"]) +
                       sum(j["Preço (C$)"] for j in time["Atacantes"]) + time["Técnico"]["Preço (C$)"])

        if custo_total <= saldo:
            return time

# Gerar quatro times
times = [criar_time(df, saldo) for _ in range(4)]

dados_saida = []
for i, time in enumerate(times, 1):
    # Encontrar o capitão (maior média de pontuação)
    todos_jogadores = []
    for pos, jgs in time.items():
        if isinstance(jgs, list):
            todos_jogadores.extend(jgs)
        else:
            todos_jogadores.append(jgs)
    capitao = max(todos_jogadores, key=lambda x: x["Média Pontuação"])

    for posicao, jogadores in time.items():
        if isinstance(jogadores, list):
            for jogador in jogadores:
                apelido = jogador["Apelido"] + (" (Capitão)" if jogador["Apelido"] == capitao["Apelido"] else "")
                dados_saida.append([f"Time {i}", posicao, apelido, jogador["Preço (C$)"], jogador["Média Pontuação"], jogador["Última Pontuação"]])
        else:
            apelido = jogadores["Apelido"] + (" (Capitão)" if jogadores["Apelido"] == capitao["Apelido"] else "")
            dados_saida.append([f"Time {i}", posicao, apelido, jogadores["Preço (C$)"], jogadores["Média Pontuação"], jogadores["Última Pontuação"]])

saida_df = pd.DataFrame(dados_saida, columns=["Time", "Posição", "Apelido", "Preço (C$)", "Média Pontuação", "Última Pontuação"])
arquivo_saida = "times_recomendados_v6.xlsx"
saida_df.to_excel(arquivo_saida, index=False)

# Melhorar layout do Excel
wb = load_workbook(arquivo_saida)
ws = wb.active

header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill("solid", fgColor="4F81BD")
alignment = Alignment(horizontal="center")

for col in ws.iter_cols(min_row=1, max_row=1):
    for cell in col:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = alignment

# Ajustar largura das colunas
for column_cells in ws.columns:
    max_length = max(len(str(cell.value)) for cell in column_cells)
    ws.column_dimensions[column_cells[0].column_letter].width = max_length + 2

wb.save(arquivo_saida)
print(f"Arquivo '{arquivo_saida}' gerado com layout melhorado!")

# Mostrar os melhores jogadores
print("\nTop 10 jogadores com melhor média de pontuação:")
print(df[["Apelido", "Clube", "Posição", "Média Pontuação"]].head(10))

print("\nTop 10 jogadores com maior chance de valorização:")
df_valorizacao = df.sort_values(by="Variação de Preço", ascending=False)
print(df_valorizacao[["Apelido", "Clube", "Posição", "Variação de Preço"]].head(10))
